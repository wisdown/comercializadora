# core/services/purchase_export_service.py

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.db.models import Sum
from django.utils.dateparse import parse_date

from core.models import Compra, CompraDetalle


class PurchaseExportService:

    @staticmethod
    def filtrar_compras(params):
        """Reutiliza filtros igual que el dashboard."""
        qs = Compra.objects.all()

        proveedor_id = params.get("proveedor_id")
        bodega_id = params.get("bodega_id")
        estado = params.get("estado")
        fecha_desde = params.get("fecha_desde")
        fecha_hasta = params.get("fecha_hasta")

        if proveedor_id:
            qs = qs.filter(proveedor_id=proveedor_id)

        if bodega_id:
            qs = qs.filter(bodega_id=bodega_id)

        if estado:
            qs = qs.filter(estado=estado.upper())

        if fecha_desde:
            d = parse_date(fecha_desde)
            if d:
                qs = qs.filter(fecha__date__gte=d)

        if fecha_hasta:
            h = parse_date(fecha_hasta)
            if h:
                qs = qs.filter(fecha__date__lte=h)

        return qs.order_by("-fecha")

    @staticmethod
    def generar_excel(qs):
        """Genera archivo XLSX con resumen y detalles."""
        wb = openpyxl.Workbook()

        # ------------------ HOJA 1: RESUMEN ------------------
        ws1 = wb.active
        ws1.title = "Resumen de Compras"

        headers = [
            "ID Compra",
            "Proveedor",
            "Bodega",
            "Fecha",
            "Documento",
            "Estado",
            "Total",
        ]
        ws1.append(headers)

        # Negritas
        for col in range(1, len(headers) + 1):
            ws1.cell(row=1, column=col).font = Font(bold=True)

        for compra in qs:
            ws1.append(
                [
                    compra.id,
                    compra.proveedor.nombre,
                    compra.bodega.nombre,
                    compra.fecha.strftime("%Y-%m-%d"),
                    compra.no_documento,
                    compra.estado,
                    float(compra.total),
                ]
            )

        # Ajuste de ancho automático
        for column in ws1.columns:
            length = max(len(str(cell.value)) for cell in column)
            ws1.column_dimensions[get_column_letter(column[0].column)].width = (
                length + 2
            )

        # ------------------ HOJA 2: DETALLE ------------------
        ws2 = wb.create_sheet("Detalle de Compras")

        headers_det = [
            "ID Compra",
            "Producto",
            "Cantidad",
            "Costo Unit.",
            "Subtotal",
        ]
        ws2.append(headers_det)
        for col in range(1, len(headers_det) + 1):
            ws2.cell(row=1, column=col).font = Font(bold=True)

        detalles = CompraDetalle.objects.filter(compra__in=qs)

        for det in detalles:
            ws2.append(
                [
                    det.compra_id,
                    det.producto.nombre,
                    float(det.cantidad),
                    float(det.costo_unit),
                    float(det.subtotal),
                ]
            )

        for column in ws2.columns:
            length = max(len(str(cell.value)) for cell in column)
            ws2.column_dimensions[get_column_letter(column[0].column)].width = (
                length + 2
            )

        return wb
